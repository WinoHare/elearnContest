import csv
import math
from openpyxl import Workbook
from openpyxl.styles import Font, Border, Side
import matplotlib.pyplot as plt
import pdfkit
from jinja2 import Environment, FileSystemLoader


class DataSet:
    def __init__(self, file_name):
        self.is_empty = False
        self.error_massage = ''
        self.file_name = file_name
        self.vacancies_objects = self.csv_reader()

    def csv_reader(self) -> tuple or str:
        is_header = True
        vacancies = []
        header = []
        header_length = 0
        with open(self.file_name, encoding='utf-8') as csv_file:
            file = csv.reader(csv_file)
            for row in file:
                if is_header:
                    is_header = False
                    header = row
                    header_length = len(header)
                    continue
                if self.is_correct_line(row, header_length):
                    vacancies.append(row)
        if len(header) == 0:
            self.is_empty = True
            self.error_massage = 'Пустой файл'
        elif len(vacancies) == 0:
            self.is_empty = True
            self.error_massage = 'Нет данных'
        return self.formatter(header, vacancies)

    def is_correct_line(self, line: list, header_length: int) -> bool:
        return '' not in line and len(line) == header_length

    def formatter(self, header: list, rows: list) -> list:
        vacancies = []
        for row in rows:
            vacancies_dict = {}
            for i in range(len(header)):
                vacancies_dict[header[i]] = row[i]
            vacancies.append(Vacancy(vacancies_dict))
        return vacancies


class Vacancy:
    def __init__(self, args: dict):
        self.name = args['name'] if 'name' in args.keys() else ''
        self.salary = Salary(args)
        self.area_name = args['area_name'] if 'area_name' in args.keys() else ''
        self.published_at = int(args['published_at'][0:4]) if 'published_at' in args.keys() else ''


class Salary:
    def __init__(self, args: dict):
        self.salary_from = args['salary_from'] if 'salary_from' in args.keys() else ''
        self.salary_to = args['salary_to'] if 'salary_to' in args.keys() else ''
        self.salary_gross = args['salary_gross'] if 'salary_gross' in args.keys() else ''
        self.salary_currency = args['salary_currency'] if 'salary_currency' in args.keys() else ''
        self.average_salary = math.floor(
            ((float(self.salary_from) + float(self.salary_to)) / 2) * self.currency_to_rub[self.salary_currency])

    currency_to_rub = {
        "AZN": 35.68,
        "BYR": 23.91,
        "EUR": 59.90,
        "GEL": 21.74,
        "KGS": 0.76,
        "KZT": 0.13,
        "RUR": 1,
        "UAH": 1.64,
        "USD": 60.66,
        "UZS": 0.0055,
    }


class InputConnect:
    def __init__(self):
        self.file_name = input('Введите название файла: ')
        self.vacancy_name = input('Введите название профессии: ')


class Statistics:
    def __init__(self, input_values):
        self.vacancies = DataSet(input_values.file_name).vacancies_objects
        self.salary_by_year = {}
        self.count_by_year = {}
        self.prof_salary_by_year = {}
        self.prof_count_by_year = {}
        self.salary_by_city = {}
        self.count_by_city = {}

    def get_statistics(self, vacancie_to_stats) -> None:
        for vacancy in self.vacancies:
            self.update_stats(vacancy.published_at, vacancy.salary.average_salary, self.salary_by_year)
            self.update_stats(vacancy.published_at, 1, self.count_by_year)
            self.update_stats(vacancy.area_name, vacancy.salary.average_salary, self.salary_by_city)
            self.update_stats(vacancy.area_name, 1, self.count_by_city)
            if vacancie_to_stats in vacancy.name:
                self.update_stats(vacancy.published_at, vacancy.salary.average_salary, self.prof_salary_by_year)
                self.update_stats(vacancy.published_at, 1, self.prof_count_by_year)

        self.get_average_salary(self.salary_by_year, self.count_by_year)
        self.get_average_salary(self.prof_salary_by_year, self.prof_count_by_year)
        self.get_average_salary(self.salary_by_city, self.count_by_city)
        self.get_percentage_of_total(self.count_by_city, len(self.vacancies))

        self.get_cities_with_enough_count()
        self.sort_statistics()
        self.print_stats()

    def update_stats(self, key: str, value: int or float, stats: dict) -> None:
        if key in stats.keys():
            stats[key] += value
        else:
            stats[key] = value

    def get_average_salary(self, salary_stats: dict, count_stats: dict) -> None:
        for key in count_stats.keys():
            salary_stats[key] = math.floor(salary_stats[key] / count_stats[key])

    def get_percentage_of_total(self, stats: dict, count: int) -> None:
        for key in stats.keys():
            stats[key] = round(stats[key] / count, 4)

    def get_cities_with_enough_count(self) -> None:
        new_salary_by_city = {}
        new_count_by_city = {}
        for key, value in self.count_by_city.items():
            if value * 100 < 1:
                continue
            new_salary_by_city[key] = self.salary_by_city[key]
            new_count_by_city[key] = value
        self.salary_by_city = new_salary_by_city
        self.count_by_city = new_count_by_city

    def sort_statistics(self) -> None:
        self.prof_salary_by_year = self.prof_salary_by_year if len(self.prof_salary_by_year) != 0 else {2022: 0}
        self.prof_count_by_year = self.prof_count_by_year if len(self.prof_count_by_year) != 0 else {2022: 0}
        self.salary_by_city = dict(sorted(self.salary_by_city.items(), key=lambda x: x[1], reverse=True)[:10])
        self.count_by_city = dict(sorted(self.count_by_city.items(), key=lambda x: x[1], reverse=True)[:10])

    def get_percent_of_other_cities(self) -> None:
        others_percent = 1
        for city_percent in self.count_by_city.values():
            others_percent -= city_percent
        self.count_by_city['Другие'] = others_percent

    def print_stats(self) -> None:
        print('Динамика уровня зарплат по годам:', self.salary_by_year)
        print('Динамика количества вакансий по годам:', self.count_by_year)
        print('Динамика уровня зарплат по годам для выбранной профессии:', self.prof_salary_by_year)
        print('Динамика количества вакансий по годам для выбранной профессии:', self.prof_count_by_year)
        print('Уровень зарплат по городам (в порядке убывания):', self.salary_by_city)
        print('Доля вакансий по городам (в порядке убывания):', self.count_by_city)


class GraphsCreator:
    def __init__(self, stats, vacancie_name):
        self.generate_image(stats, vacancie_name)

    def generate_image(self, stats, vacancie_name):
        plt.subplots(figsize=(10, 7))
        plt.grid(True)

        self.create_salary_by_year_plot(stats, vacancie_name)
        self.create_count_by_year_plot(stats, vacancie_name)
        self.create_salary_by_city_plot(stats)
        self.create_count_by_city_plot(stats)

        plt.subplots_adjust(wspace=0.5, hspace=0.5)
        plt.savefig('graph.png')

    def create_salary_by_year_plot(self, stats, vacancie_name):
        first = plt.subplot(221)
        plt.tick_params(axis='x', which='major', labelsize=8, rotation=90)
        plt.tick_params(axis='y', which='major', labelsize=8)
        first.bar(list(map(lambda y: y - 0.2, stats.salary_by_year.keys())),
                  stats.salary_by_year.values(), width=0.4,
                  label='Средняя з/п')
        first.bar(list(map(lambda y: y + 0.2, stats.prof_salary_by_year.keys())),
                  stats.prof_salary_by_year.values(), width=0.4,
                  label=f'З/п {vacancie_name}')
        plt.legend(fontsize=8)
        plt.title('Уровень зарплат по годам', fontsize=12)

    def create_count_by_year_plot(self, stats, vacancie_name):
        second = plt.subplot(222)
        plt.tick_params(axis='x', which='major', labelsize=8, rotation=90)
        plt.tick_params(axis='y', which='major', labelsize=8)
        second.bar(list(map(lambda y: y - 0.2, stats.count_by_year.keys())),
                   stats.count_by_year.values(), width=0.4,
                   label='Количество вакансий')
        second.bar(list(map(lambda y: y + 0.2, stats.prof_count_by_year.keys())),
                   stats.prof_count_by_year.values(), width=0.4,
                   label=f'Количество вакансий {vacancie_name}')
        plt.legend(fontsize=8)
        plt.title('Количество вакансий по годам', fontsize=12)

    def create_salary_by_city_plot(self, stats):
        third = plt.subplot(223)
        plt.tick_params(axis='x', which='major', labelsize=8)
        plt.tick_params(axis='y', which='major', labelsize=6)
        third.barh(list(reversed(stats.salary_by_city.keys())),
                   list(reversed(stats.salary_by_city.values())))
        plt.title('Уровень зарплат по городам', fontsize=12)

    def create_count_by_city_plot(self, stats):
        stats.get_percent_of_other_cities()
        fourth = plt.subplot(224)
        plt.rc('xtick', labelsize=6)
        fourth.pie(list(map(lambda c: round(c * 100, 2), stats.count_by_city.values())),
                   labels=stats.count_by_city.keys(),
                   colors=['r', 'g', 'b', 'm', 'y', 'c', 'orange', 'darkblue', 'pink', 'sienna', 'grey'])
        plt.title('Доля вакансий по городам', fontsize=12)


class ExcelCreator:
    def __init__(self, stats: Statistics, vacancie_name):
        self.workbook = self.initialize_workbook(stats, vacancie_name)

    def initialize_workbook(self, stats: Statistics, vacancie_name):
        workbook, years_sheet, cities_sheet = self.create_workbook(vacancie_name)
        self.add_stats_to_excel(stats, years_sheet, cities_sheet)
        self.set_sheets_settings(stats, years_sheet, cities_sheet)
        workbook.save('report.xlsx')
        return workbook

    def create_workbook(self, vacancie_name) -> tuple:
        workbook = Workbook()

        years_sheet = workbook.active
        years_sheet.title = 'Статистика по годам'
        cities_sheet = workbook.create_sheet('Статистика по городам')

        years_sheet.append(
            ['Год', 'Средняя зарплата', f'Средняя зарплата - {vacancie_name}',
             'Количество вакансий', f'Количество вакансий - {vacancie_name}'])
        cities_sheet.append(['Город', 'Уровень зарплат', '', 'Город', 'Доля Вакансий'])

        return workbook, years_sheet, cities_sheet

    def add_stats_to_excel(self, stats: Statistics, years_sheet, cities_sheet):
        for year in stats.salary_by_year.keys():
            years_sheet.append([year,
                                stats.salary_by_year[year],
                                stats.prof_salary_by_year[year],
                                stats.count_by_year[year],
                                stats.prof_salary_by_year[year]])

        for city in stats.salary_by_city.keys():
            cities_sheet.append([city, stats.salary_by_city[city]])

        for i, city in enumerate(stats.count_by_city.keys(), 2):
            cities_sheet[f'D{i}'].value = city
            cities_sheet[f'E{i}'].value = f'{round(stats.count_by_city[city] * 100, 2)}%'

    def set_sheets_settings(self, stats: Statistics, years_sheet, cities_sheet) -> None:
        used_columns = ['A', 'B', 'C', 'D', 'E']
        for i in used_columns:
            years_sheet[f'{i}1'].font = Font(bold=True)
            cities_sheet[f'{i}1'].font = Font(bold=True)
            years_sheet.column_dimensions[i].width = max(map(lambda x: len(str(x.value)), years_sheet[i])) + 1
            cities_sheet.column_dimensions[i].width = max(
                map(lambda x: len(str(x.value)), cities_sheet[i])) + 1

        thins = Side(border_style="thin")
        for column in used_columns:
            for row in range(1, len(stats.salary_by_year.keys()) + 2):
                years_sheet[f'{column}{row}'].border = Border(top=thins, bottom=thins, left=thins, right=thins)

        for column in used_columns:
            for row in range(1, len(stats.salary_by_city.keys()) + 2):
                if column == 'C':
                    break
                cities_sheet[f'{column}{row}'].border = Border(top=thins, bottom=thins, left=thins, right=thins)


class PdfCreator:
    def __init__(self, vacancie_name, workbook, years_sheet_rows, cities_sheet_rows):
        config = pdfkit.configuration(wkhtmltopdf=r'D:\Program Files\wkhtmltopdf\bin\wkhtmltopdf.exe')
        options = {'enable-local-file-access': None}
        pdfkit.from_string(self.fill_pdf_template(vacancie_name,
                                                  workbook["Статистика по годам"],
                                                  workbook["Статистика по городам"],
                                                  years_sheet_rows,
                                                  cities_sheet_rows),
                           'report.pdf', configuration=config, options=options)

    def fill_pdf_template(self, vacancie_name, years_sheet, cities_sheet, years_sheet_rows, cities_sheet_rows):
        env = Environment(loader=FileSystemLoader('.'))
        template = env.get_template('pdf_template.html')
        pdf_template = template.render({'vacancie_name': vacancie_name,
                                        'years_table': self.create_html_table(years_sheet, years_sheet_rows),
                                        'cities_table_first': self.create_html_table(cities_sheet,
                                                                                     cities_sheet_rows, last_column=2),
                                        'cities_table_second': self.create_html_table(cities_sheet,
                                                                                      cities_sheet_rows, 4)})
        return pdf_template

    def create_html_table(self, ws, rows_count, first_column = 1, last_column=5):
        html = ''
        is_first = True
        for row in ws.iter_rows(min_row=1, min_col=first_column, max_col=last_column, max_row=rows_count + 1):
            html += '<tr>'
            for cell in row:
                if cell.value is None or cell.value == '':
                    html += "<td>" + '   ' + "</td>"
                else:
                    html += '<td><b>' + str(cell.value) + '</b></td>' if is_first else '<td>' + str(
                        cell.value) + '</td>'
            html += '</tr>'
            is_first = False
        return html


class Report:
    def __init__(self):
        self.input_values = InputConnect()
        self.statistics = Statistics(self.input_values)
        self.statistics.get_statistics(self.input_values.vacancy_name)

        workbook = ExcelCreator(self.statistics, self.input_values.vacancy_name)
        graph_image = GraphsCreator(self.statistics, self.input_values.vacancy_name)
        pdf = PdfCreator(self.input_values.vacancy_name, workbook.workbook, len(self.statistics.salary_by_year.keys()),
                         len(self.statistics.salary_by_city.keys()))


report = Report()
